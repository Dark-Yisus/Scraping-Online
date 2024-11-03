from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
import logging
from pymongo import MongoClient, UpdateOne
from bson import json_util
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# Initialize Flask app
app = Flask(__name__)

# Configure CORS with specific origins and methods
CORS(app, resources={
    r"/mercadolibre": {
        "origins": ["http://mercado-scraping.shop", "https://mercado-scraping.shop"],
        "methods": ["POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    },
    r"/descargarExcel": {
        "origins": ["http://mercado-scraping.shop", "https://mercado-scraping.shop"],
        "methods": ["POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
try:
    client = MongoClient('mongodb://localhost:27017/')
    db = client['mercadolibre_db']
    collection = db['productos']
    logger.info("Successfully connected to MongoDB")
except Exception as e:
    logger.error(f"Failed to connect to MongoDB: {e}")

# CORS middleware
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', 'https://mercado-scraping.shop')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

def extract_product_details(url_producto, headers):
    """
    Extracts product details from a MercadoLibre product URL
    """
    try:
        r_producto = None
        for _ in range(3):  # Try up to 3 times
            r_producto = requests.get(url_producto, headers=headers, timeout=10)
            if r_producto.status_code == 200:
                break

        if not r_producto or r_producto.status_code != 200:
            logger.warning(f"Failed to fetch product page: {url_producto}")
            return None

        soup_producto = BeautifulSoup(r_producto.content, 'html.parser')

        # Extract seller information
        vendedor_tag = soup_producto.find('div', class_='ui-pdp-seller__header__title')
        vendedor = vendedor_tag.get_text(strip=True) if vendedor_tag else 'N/A'

        # Extract prices
        precio_original = 'N/A'
        precio_con_descuento = 'N/A'
        descuento = 'N/A'

        precio_original_tag = soup_producto.find('s', class_='andes-money-amount ui-pdp-price__part ui-pdp-price__original-value andes-money-amount--previous andes-money-amount--cents-superscript andes-money-amount--compact')
        if precio_original_tag:
            precio_original = precio_original_tag.text.strip()

        precio_con_descuento_tag = soup_producto.find('div', class_='ui-pdp-price__second-line')
        if precio_con_descuento_tag:
            precio_descuento = precio_con_descuento_tag.find('span', class_='andes-money-amount__fraction')
            if precio_descuento:
                precio_con_descuento = precio_descuento.text.strip()

        # Extract discount percentage
        descuento_tag = soup_producto.find('span', class_='ui-pdp-price__second-line__label')
        if descuento_tag:
            descuento = descuento_tag.text.strip()

        # Extract payment installments
        cuotas_pago = 'N/A'
        cuotas_container = soup_producto.find('div', class_='ui-pdp-payment')
        if cuotas_container:
            cuotas_text = cuotas_container.get_text(strip=True)
            cuotas_match = re.search(r'(\d+)x\s*(\$[\d,.]+)\s*(sin interés|con interés)?', cuotas_text)
            if cuotas_match:
                cuotas_pago = f"{cuotas_match.group(1)}x {cuotas_match.group(2)} {cuotas_match.group(3) or ''}"

        # Extract interest-free months
        meses_sin_intereses = 'N/A'
        meses_sin_intereses_tag = soup_producto.find('span', class_='ui-pdp-color--GREEN')
        if meses_sin_intereses_tag:
            meses_sin_intereses = meses_sin_intereses_tag.get_text(strip=True)

        # Extract shipping information
        envio_tag = soup_producto.find('p', class_='ui-pdp-color--BLACK ui-pdp-family--REGULAR ui-pdp-media__title')
        envio = envio_tag.get_text(strip=True) if envio_tag else 'N/A'

        # Extract number of items sold
        cantidad_vendida = 'N/A'
        cantidad_tag = soup_producto.find('span', class_='ui-pdp-subtitle')
        if cantidad_tag:
            cantidad_match = re.search(r'(\d+)\s+vendidos?', cantidad_tag.text)
            if cantidad_match:
                cantidad_vendida = cantidad_match.group(1)

        # Extract image URL
        imagen_tag = soup_producto.find('img', class_='ui-pdp-image ui-pdp-gallery__figure__image')
        imagen = imagen_tag['src'] if imagen_tag and 'src' in imagen_tag.attrs else 'N/A'

        return {
            'vendedor': vendedor,
            'precio_original': precio_original,
            'precio_con_descuento': precio_con_descuento,
            'descuento': descuento,
            'cuotas': cuotas_pago,
            'meses_sin_intereses': meses_sin_intereses,
            'envios': envio,
            'cantidad_vendida': cantidad_vendida,
            'imagenes': imagen
        }
    except Exception as e:
        logger.error(f"Error processing product page {url_producto}: {e}")
        return None

def search_product_api(producto):
    """
    Searches for products using the MercadoLibre API
    """
    try:
        url = f'https://api.mercadolibre.com/sites/MLM/search?q={producto}'
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"Error searching product in API: {e}")
        return None

def save_products_to_db(productos_data):
    """
    Saves or updates products in MongoDB
    """
    if not productos_data:
        return

    try:
        operations = []
        for producto in productos_data:
            filter_query = {"url_producto": producto["url_producto"]}
            update_query = {"$set": producto}
            operations.append(UpdateOne(filter_query, update_query, upsert=True))
        
        result = collection.bulk_write(operations)
        logger.info(f"Saved/updated {result.upserted_count + result.modified_count} products in database")
    except Exception as e:
        logger.error(f"Error inserting/updating products in MongoDB: {e}")

@app.route('/mercadolibre', methods=['POST', 'OPTIONS'])
def search_products():
    """
    Main endpoint for product search
    """
    if request.method == 'OPTIONS':
        return '', 204
        
    try:
        data = request.get_json()
        if not data or 'producto' not in data:
            return jsonify({'error': 'No se proporcionó ningún producto'}), 400
        
        producto = data['producto']
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        start_time = datetime.now()
        productos_a_guardar = []

        resultado_api = search_product_api(producto)
        if not resultado_api:
            return jsonify({'error': 'Error searching products'}), 500

        for item in resultado_api['results']:
            url_producto = item['permalink']
            detalles = extract_product_details(url_producto, headers)

            if detalles:
                producto_data = {
                    "titulo": item['title'],
                    "url_producto": url_producto,
                    **detalles,
                    "fecha_extraccion": datetime.now()
                }
                productos_a_guardar.append(producto_data)

        save_products_to_db(productos_a_guardar)

        return jsonify({
            "datos": productos_a_guardar,
            "num_products": len(productos_a_guardar),
            "processing_time": (datetime.now() - start_time).total_seconds()
        })

    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/descargarExcel', methods=['POST', 'OPTIONS'])
def download_excel():
    """
    Endpoint to download results in Excel format
    """
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = json_util.loads(request.form.get('data'))
        
        # Create DataFrame
        df = pd.DataFrame(data['datos'])
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos MercadoLibre"
        
        # Add headers
        headers = list(df.columns)
        ws.append(headers)
        
        # Add data
        for _, row in df.iterrows():
            ws.append(row.tolist())
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel to memory
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='productos_mercadolibre.xlsx'
        )
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(
        host='0.0.0.0', 
        port=5000, 
        ssl_context=('cert.pem', 'key.pem'),
        debug=False
    )